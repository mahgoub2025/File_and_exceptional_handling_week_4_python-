import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from PIL import Image


def handle_text_file(filename):
    """Read, modify, write, or append text files"""
    try:
        with open(filename, "r", encoding="utf-8") as f:
            content = f.read()

        print("\nChoose a modification option:")
        print("1. Convert to UPPERCASE")
        print("2. Convert to lowercase")
        print("3. Add line numbers")
        print("4. Append custom text (new file)")
        print("5. Append custom text (same file)")

        choice = input("Enter choice (1-5): ").strip()

        if choice == "1":
            modified = content.upper()
            output_file = "modified_" + os.path.basename(filename)
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(modified)
            print(f"Modified text file saved as: {output_file}")

        elif choice == "2":
            modified = content.lower()
            output_file = "modified_" + os.path.basename(filename)
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(modified)
            print(f"Modified text file saved as: {output_file}")

        elif choice == "3":
            lines = content.splitlines()
            modified = "\n".join(f"{i+1}: {line}" for i, line in enumerate(lines))
            output_file = "modified_" + os.path.basename(filename)
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(modified)
            print(f"Modified text file saved as: {output_file}")

        elif choice == "4":
            extra = input("Enter text to append (saved to a new file): ")
            modified = content + "\n" + extra
            output_file = "modified_" + os.path.basename(filename)
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(modified)
            print(f"Modified text file saved as: {output_file}")

        elif choice == "5":
            extra = input("Enter text to append (added directly to same file): ")
            with open(filename, "a", encoding="utf-8") as f:
                f.write("\n" + extra)
            print(f"Text appended directly to: {filename}")

        else:
            print("Invalid choice, no modification applied.")

    except Exception as e:
        print(f"Error handling text file: {e}")


def handle_excel_file(filename):
    """Read, modify, or append Excel files"""
    try:
        wb = load_workbook(filename)
        sheet = wb.active

        print("\nChoose Excel operation:")
        print("1. Append new row (saved as new file)")
        print("2. Append new row (to same file)")

        choice = input("Enter choice (1/2): ").strip()

        new_row = ["Appended", "Row", "Data"]

        if choice == "1":
            sheet.append(new_row)
            output_file = "modified_" + os.path.basename(filename)
            wb.save(output_file)
            print(f"Modified Excel file saved as: {output_file}")

        elif choice == "2":
            sheet.append(new_row)
            wb.save(filename)
            print(f"New row appended directly to: {filename}")

        else:
            print("Invalid choice, no modification applied.")

    except Exception as e:
        print(f"Error handling Excel file: {e}")


def handle_pdf_file(filename):
    """Extract or append PDF text"""
    try:
        reader = PdfReader(filename)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

        print("\nChoose PDF operation:")
        print("1. Save extracted text to new file")
        print("2. Append extracted text to existing .txt file")

        choice = input("Enter choice (1/2): ").strip()

        if choice == "1":
            output_file = os.path.splitext(os.path.basename(filename))[0] + "_extracted.txt"
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(text)
            print(f"Extracted text saved as: {output_file}")

        elif choice == "2":
            target = input("Enter .txt filename to append into: ").strip()
            with open(target, "a", encoding="utf-8") as f:
                f.write(f"\n--- Extracted from {filename} ---\n")
                f.write(text)
            print(f"PDF text appended into: {target}")

        else:
            print("Invalid choice, no modification applied.")

    except Exception as e:
        print(f"Error handling PDF file: {e}")


def handle_image_file(filename):
    """Read and display image metadata"""
    try:
        with Image.open(filename) as img:
            print(f"Image format: {img.format}")
            print(f"Image size: {img.size}")
            print(f"Image mode: {img.mode}")

    except Exception as e:
        print(f"Error handling image file: {e}")


def read_file_only(filename):
    """Smart read: text, excel, pdf, or image"""
    try:
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".txt":
            with open(filename, "r", encoding="utf-8") as f:
                print("\n=== File Content (Text) ===")
                print(f.read())

        elif ext == ".xlsx":
            wb = load_workbook(filename)
            sheet = wb.active
            print("\n=== File Content (Excel Preview) ===")
            for row in sheet.iter_rows(values_only=True):
                print(row)

        elif ext == ".pdf":
            reader = PdfReader(filename)
            print("\n=== File Content (PDF Text) ===")
            for i, page in enumerate(reader.pages, start=1):
                text = page.extract_text()
                if text:
                    print(f"\n--- Page {i} ---\n{text}")

        elif ext in [".jpg", ".jpeg", ".png"]:
            with Image.open(filename) as img:
                print("\n=== File Info (Image) ===")
                print(f"Format: {img.format}")
                print(f"Size: {img.size}")
                print(f"Mode: {img.mode}")

        else:
            print("Unsupported file type for preview.")

    except Exception as e:
        print(f"Error reading file: {e}")


def main():
    while True:
        print("\nFile Handling Menu")
        print("1. Text File (.txt)")
        print("2. Excel File (.xlsx)")
        print("3. PDF File (.pdf)")
        print("4. Image File (.jpg/.png)")
        print("5. Exit")
        print("6. Read any file (display only)")

        choice = input("Select a file type (1-6): ").strip()

        if choice == "5":
            print("Exiting program.")
            break

        elif choice == "6":
            filename = input("Enter the filename (with extension): ").strip()
            if not os.path.exists(filename):
                print(f"File '{filename}' not found!")
                continue
            read_file_only(filename)
            continue

        filename = input("Enter the filename (with extension): ").strip()

        if not os.path.exists(filename):
            print(f"File '{filename}' not found!")
            continue

        if choice == "1":
            handle_text_file(filename)
        elif choice == "2":
            handle_excel_file(filename)
        elif choice == "3":
            handle_pdf_file(filename)
        elif choice == "4":
            handle_image_file(filename)
        else:
            print("Invalid choice. Try again.")


if __name__ == "__main__":
    main()
